import os
from itertools import product
from uuid import uuid4
from collections import namedtuple

import stackprinter
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, NamedStyle, PatternFill

TEMP_CC_DataInfo = namedtuple("TEMP_CC_DataInfo", "top btm env")
ILLU_CC_DataInfo = namedtuple("ILLU_CC_DataInfo", "wave standard_pointses channel_pointses")
SAMPLE_TITLES = ("索引", "日期", "标签", "控制板程序版本", "控制板芯片ID", "通道", "项目", "方法", "波长")
CORRECT_HEADS = ("通道", "标点")

ODD_STYLE = NamedStyle(name="odd")
ODD_STYLE.fill = PatternFill(start_color="bedede", end_color="bedede", fill_type="solid")
ODD_STYLE.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
EVEN_STYLE = NamedStyle(name="even")
EVEN_STYLE.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
EVEN_STYLE.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)

CC_STYLE = NamedStyle(name="cc")
CC_STYLE.fill = PatternFill(start_color="3bee35", end_color="3bee35", fill_type="solid")
CC_STYLE.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
OO_STYLE = NamedStyle(name="oo")
OO_STYLE.fill = PatternFill(start_color="f9a0a0", end_color="f9a0a0", fill_type="solid")
OO_STYLE.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)


DEFAULT_CC_DATA = []
for t_idx, title in enumerate((610, 550, 405)):
    standard_pointses = []
    channel_pointses = []
    cn = 6 if t_idx != 2 else 1
    for row_idx in range(cn):
        channel_pointses.append(tuple(0 for column_idx in range(6)))
        standard_pointses.append(tuple(0 for column_idx in range(6)))
    DEFAULT_CC_DATA.append(ILLU_CC_DataInfo(wave=title, standard_pointses=standard_pointses, channel_pointses=tuple(channel_pointses)))
DEFAULT_CC_DATA.append(TEMP_CC_DataInfo(top=0, btm=0, env=0))
DEFAULT_CC_DATA = tuple(DEFAULT_CC_DATA)


def dump_CC(data, file_path):
    wb = Workbook(write_only=True)
    sheets = [wb.create_sheet(f"{title}") for title in (610, 550, 405, "温度")]
    # build frame
    for idx, sheet in enumerate(sheets[:3]):
        # https://www.johndcook.com/wavelength_to_RGB.html
        sheet.sheet_properties.tabColor = ("ff9b00", "a3ff00", "8200c8")[idx]
        cells = [WriteOnlyCell(sheet, value=(sheet.title, "测试点-1", "测试点-2", "测试点-3", "测试点-4", "测试点-5", "测试点-6")[i]) for i in range(7)]
        sheet.append(cells)
    sheet = sheets[-1]
    cells = [WriteOnlyCell(sheet, value=("上加热体", "下加热体", "环境")[i]) for i in range(3)]
    sheet.append(cells)
    # insert data
    try:
        for d in data:
            if isinstance(d, ILLU_CC_DataInfo):
                sheet = wb[f"{d.wave}"]
                for row_idx, sc_ps in enumerate(zip(d.standard_pointses, d.channel_pointses)):
                    sps, cps = sc_ps
                    cells_s = [WriteOnlyCell(sheet, value=f"通道-{row_idx + 1}-标准值")] + [WriteOnlyCell(sheet, value=sp) for sp in sps]
                    sheet.append(cells_s)
                    cells_c = [WriteOnlyCell(sheet, value=f"通道-{row_idx + 1}-实际值")] + [WriteOnlyCell(sheet, value=cp) for cp in cps]
                    sheet.append(cells_c)
            elif isinstance(d, TEMP_CC_DataInfo):
                sheet = wb["温度"]
                cells = (WriteOnlyCell(sheet, value=d.top), WriteOnlyCell(sheet, value=d.btm), WriteOnlyCell(sheet, value=d.env))
                sheet.append(cells)
        wb.save(find_new_file(file_path))
        return True
    except Exception:
        logger.error(f"save data to xlsx failed\n{stackprinter.format()}")


def load_CC(file_path):
    if not os.path.isfile(file_path):
        return None
    try:
        wb = load_workbook(file_path, read_only=True)
        data = []
        for t_idx, title in enumerate((610, 550, 405)):
            sheet_name = f"{title}"
            if sheet_name not in wb.sheetnames:
                return DEFAULT_CC_DATA
            sheet = wb[sheet_name]
            standard_pointses = []
            channel_pointses = []
            cn = 6 if t_idx != 2 else 1
            for row_idx in range(cn):
                standard_pointses.append(tuple(sheet.cell(column=column_idx + 2, row=2 + 2 * row_idx).value for column_idx in range(6)))
                channel_pointses.append(tuple(sheet.cell(column=column_idx + 2, row=3 + 2 * row_idx).value for column_idx in range(6)))
            data.append(ILLU_CC_DataInfo(wave=title, standard_pointses=tuple(standard_pointses), channel_pointses=tuple(channel_pointses)))
        if "温度" in wb.sheetnames:
            sheet = wb["温度"]
            data.append(TEMP_CC_DataInfo(top=sheet.cell(column=1, row=2).value, btm=sheet.cell(column=2, row=2).value, env=sheet.cell(column=3, row=2).value))
        else:
            data.append(TEMP_CC_DataInfo(top=0, btm=0, env=0))
        return tuple(data)
    except Exception:
        logger.error(f"load data from xlsx failed\n{stackprinter.format()}")
        return None


def cc_get_pointes_info(cc, channel, wave):
    for data in cc:
        if isinstance(data, ILLU_CC_DataInfo) and data.wave == wave:
            return data.channel_pointses[channel - 1], data.standard_pointses[channel - 1]
    return None, None


def check_file_permission(file_path):
    if not os.path.isfile(file_path):
        return True
    old_path = file_path
    new_path = f"{file_path}{uuid4().hex}"
    while os.path.isfile(new_path):
        new_path = f"{file_path}{uuid4().hex}"
    try:
        os.rename(old_path, new_path)
    except PermissionError:
        return False
    else:
        return True
    finally:
        if os.path.isfile(new_path):
            os.rename(new_path, old_path)


def find_new_file(file_path):
    if not check_file_permission(file_path):
        logger.warning(f"file_path could not be renamed | {file_path}")
        attemp = 1
        dir_path, file_name = os.path.split(file_path)
        new_file_name = f"{os.path.splitext(file_name)[0]}_{attemp}{os.path.splitext(file_name)[1]}"
        while os.path.isfile(new_file_name):
            new_file_name = f"{os.path.splitext(file_name)[0]}_{attemp}{os.path.splitext(file_name)[1]}"
            attemp += 1
        file_path = os.path.join(dir_path, new_file_name)
        logger.warning(f"new file_path become | {file_path}")
    return file_path


def dump_sample(sample_iter, file_path, title=None):
    try:
        wb = Workbook(write_only=True)
        wb.add_named_style(ODD_STYLE)
        wb.add_named_style(EVEN_STYLE)
        wb.add_named_style(CC_STYLE)
        wb.add_named_style(OO_STYLE)
        # title
        if title is None:
            title = f"Sheet{len(wb.sheetnames) + 1}"
        # sheet
        sheet = wb.create_sheet(f"{title}")
        sheet.freeze_panes = "G2"
        sheet.column_dimensions["A"].width = 6.25
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 27
        sheet.column_dimensions["D"].width = 0.01
        sheet.column_dimensions["E"].width = 0.01
        sheet.column_dimensions["F"].width = 5
        sheet.column_dimensions["G"].width = 10
        sheet.column_dimensions["H"].width = 9
        sheet.column_dimensions["I"].width = 5
        # head
        cells = [WriteOnlyCell(sheet, value=SAMPLE_TITLES[i]) for i in range(len(SAMPLE_TITLES))] + [WriteOnlyCell(sheet, value=i) for i in range(1, 121)]
        for cell in cells:
            cell.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
        sheet.append(cells)
        # body
        last_label_id = None
        label_cnt = 0
        for si in sample_iter:
            cells = [
                WriteOnlyCell(sheet, value=si.label_id),
                WriteOnlyCell(sheet, value=si.label_datetimne),
                WriteOnlyCell(sheet, value=si.label_name),
                WriteOnlyCell(sheet, value=si.label_version),
                WriteOnlyCell(sheet, value=si.label_device_id),
                WriteOnlyCell(sheet, value=si.sample_channel),
                WriteOnlyCell(sheet, value=si.sample_set_info),
                WriteOnlyCell(sheet, value=si.sample_method),
                WriteOnlyCell(sheet, value=si.sample_wave),
            ] + [WriteOnlyCell(sheet, value=sp) for sp in si.sample_datas]
            if si.label_id != last_label_id:
                last_label_id = si.label_id
                label_cnt += 1
            if len(si.sample_datas) / si.sample_data_total == 3:
                for cell in cells[-1 * si.sample_data_total :]:
                    cell.style = CC_STYLE
                for cell in cells[-2 * si.sample_data_total : -1 * si.sample_data_total]:
                    cell.style = OO_STYLE
            for cell in cells:
                if cell.style in ("cc", "oo"):
                    continue
                if label_cnt % 2 == 0:
                    cell.style = EVEN_STYLE
                else:
                    cell.style = ODD_STYLE
            sheet.append(cells)
        wb.save(find_new_file(file_path))
        logger.success("finish dump db to excel")
        return wb, None
    except Exception as e:
        logger.error(f"dump sample data to xlsx failed\n{stackprinter.format()}")
        return None, e


def insert_sample(sample_iter, file_path):
    sample_list = [i for i in sample_iter]
    try:
        if os.path.isfile(file_path):
            wb = load_workbook(file_path)
            new_flag = False
        else:
            wb = Workbook()
            new_flag = True
        if ODD_STYLE.name not in wb.style_names:
            wb.add_named_style(ODD_STYLE)
        if EVEN_STYLE.name not in wb.style_names:
            wb.add_named_style(EVEN_STYLE)
        # sheet
        sheet = wb.active
        if new_flag:
            sheet.freeze_panes = "G2"
            sheet.column_dimensions["A"].width = 6.25
            sheet.column_dimensions["B"].width = 30
            sheet.column_dimensions["C"].width = 27
            sheet.column_dimensions["D"].width = 0.01
            sheet.column_dimensions["E"].width = 0.01
            sheet.column_dimensions["F"].width = 5
            sheet.column_dimensions["G"].width = 10
            sheet.column_dimensions["H"].width = 9
            sheet.column_dimensions["I"].width = 5
        # head
        if new_flag:
            cells = [sheet.cell(column=i + 1, row=1, value=s) for i, s in enumerate(SAMPLE_TITLES)] + [
                sheet.cell(column=i + len(SAMPLE_TITLES), row=1, value=i) for i in range(1, 121)
            ]
            for cell in cells:
                cell.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
        # body
        last_cell = sheet["A2"]
        last_id = last_cell.value
        record_id = None
        new_style = "odd"
        if last_cell.has_style and last_cell.style == "odd":
            new_style = "even"
        sheet.insert_rows(2, len(sample_list))
        for idx, si in enumerate(sample_list):
            if si.label_id == last_id:
                logger.warning(f"not insert for same label id | {last_id} | {si}")
                continue
            cells = [
                sheet.cell(column=1, row=2 + idx, value=si.label_id),
                sheet.cell(column=2, row=2 + idx, value=si.label_datetimne),
                sheet.cell(column=3, row=2 + idx, value=si.label_name),
                sheet.cell(column=4, row=2 + idx, value=si.label_version),
                sheet.cell(column=5, row=2 + idx, value=si.label_device_id),
                sheet.cell(column=6, row=2 + idx, value=si.sample_channel),
                sheet.cell(column=7, row=2 + idx, value=si.sample_set_info),
                sheet.cell(column=8, row=2 + idx, value=si.sample_method),
                sheet.cell(column=9, row=2 + idx, value=si.sample_wave),
            ] + [sheet.cell(column=10 + i, row=2 + idx, value=sp) for i, sp in enumerate(si.sample_datas)]
            if record_id is None:
                record_id = si.label_id
            elif record_id != si.label_id:
                record_id = si.label_id
                new_style = "odd" if new_style == "even" else "even"
            for cell in cells:
                cell.style = new_style
        wb.save(find_new_file(file_path))
        logger.success("finish insert db to excel")
        return wb, None
    except Exception as e:
        logger.error(f"insert sample data to xlsx failed\n{stackprinter.format()}")
        return None, e


def dump_correct_record(correct_list, file_path):
    try:
        wb = Workbook(write_only=True)
        wb.add_named_style(ODD_STYLE)
        wb.add_named_style(EVEN_STYLE)
        wb.add_named_style(CC_STYLE)
        wb.add_named_style(OO_STYLE)
        sheet_610 = wb.create_sheet("610")
        sheet_550 = wb.create_sheet("550")
        sheet_405 = wb.create_sheet("405")

        for i in range(36):
            n = ord("C") - ord("A") + i + 1
            c = ""
            while n > 0:
                n, remainder = divmod(n - 1, 26)
                c = chr(65 + remainder) + c
            sheet_610.column_dimensions[c].width = 11
            sheet_550.column_dimensions[c].width = 11
            sheet_405.column_dimensions[c].width = 11

        for sheet in (sheet_610, sheet_550, sheet_405):
            cells = [WriteOnlyCell(sheet, value=i) for i in CORRECT_HEADS] + [
                WriteOnlyCell(sheet, value=f"{i}-{j+1}") for i, j in product(("白物质", "反应区", "OD"), list(range(12)))
            ]
            for cell in cells:
                cell.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
            sheet.append(cells)

        for cu in correct_list:
            if cu.wave == 610:
                sheet = sheet_610
            elif cu.wave == 550:
                sheet = sheet_550
            elif cu.wave == 405:
                sheet = sheet_405
            else:
                continue
            cells = [WriteOnlyCell(sheet, value=i) for i in (cu.channel, cu.stage)] + [
                WriteOnlyCell(sheet, value=i) for i in (cu.white_pd + cu.react_pd + cu.od)
            ]
            sheet.append(cells)
        wb.save(find_new_file(file_path))
        logger.success("finish dump correct flash data")
    except Exception:
        logger.error(f"dump_correct_record failed\n{stackprinter.format()}")


if __name__ == "__main__":
    file_path = r"C:\Users\Administrator\Desktop\test.xlsx"
    data = (
        ILLU_CC_DataInfo(
            wave=610,
            standard_pointses=(
                (3100, 4200, 5300, 6400, 7500, 8600),
                (3101, 4201, 5301, 6401, 7501, 8601),
                (3102, 4202, 5302, 6402, 7502, 8602),
                (3103, 4203, 5303, 6403, 7503, 8603),
                (3104, 4204, 5304, 6404, 7504, 8604),
                (3105, 4205, 5305, 6405, 7505, 8605),
            ),
            channel_pointses=(
                (1100, 2200, 3300, 4400, 5500, 6600),
                (1101, 2201, 3301, 4401, 5501, 6601),
                (1102, 2202, 3302, 4402, 5502, 6602),
                (1103, 2203, 3303, 4403, 5503, 6603),
                (1104, 2204, 3304, 4404, 5504, 6604),
                (1105, 2205, 3305, 4405, 5505, 6605),
            ),
        ),
        ILLU_CC_DataInfo(
            wave=550,
            standard_pointses=(
                (1200, 2300, 3400, 4500, 5600, 6700),
                (1201, 2301, 3401, 4501, 5601, 6701),
                (1202, 2302, 3402, 4502, 5602, 6702),
                (1203, 2303, 3403, 4503, 5603, 6703),
                (1204, 2304, 3404, 4504, 5604, 6704),
                (1205, 2305, 3405, 4505, 5605, 6705),
            ),
            channel_pointses=(
                (7100, 8200, 9300, 10400, 11500, 12600),
                (7101, 8201, 9301, 10401, 11501, 12601),
                (7102, 8202, 9302, 10402, 11502, 12602),
                (7103, 8203, 9303, 10403, 11503, 12603),
                (7104, 8204, 9304, 10404, 11504, 12604),
                (7105, 8205, 9305, 10405, 11505, 12605),
            ),
        ),
        ILLU_CC_DataInfo(wave=405, standard_pointses=((181, 172, 163, 154, 145, 136),), channel_pointses=((18100, 17200, 16300, 15400, 14500, 13600),)),
        TEMP_CC_DataInfo(top=1, btm=2, env=6),
    )
    dump_CC(data, file_path)
    rdata = load_CC(file_path)
    assert rdata == data
